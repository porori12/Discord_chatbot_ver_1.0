import discord
import datetime
import random
import os
import zipfile
import csv
import chardet
import asyncio
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


# 개인의 토큰으로 대체
# token = ''

intents = discord.Intents.all()
intents.members = True

client = discord.Client(intents=intents)

response1 = ()
response2 = ()
response3 = ()
response4 = ()
response5 = ()
card_use_date = ()


high_density_list = []
intelligent_list = []
external_people = []

meeting_fee = 0

@client.event
async def on_ready():
    print("봇 작동 중!!")
    print(client.user)
    print("=========================")

    # 봇의 상태 변경
    game = discord.Game("서류 작성")
    await client.change_presence(status=discord.Status.idle, activity=game)

@client.event
async def on_message(message):
    global meeting_fee, card_use_date

    if message.content == "고밀도 회의비 작성해줘":
        print("\n\n=========================\n고밀도 회의비 작성중......")
        print(datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        wb = Workbook()
        ws = wb.active 

        high_density_list = []


        with open('high_density_data.csv', 'r', newline='', encoding='EUC-KR') as f:
            reader = csv.reader(f)
            for row in reader:
                high_density_list.append(row)
        
        print(f"고밀도 리스트 :" , high_density_list)

        medium_border = Border(left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            bottom=Side(border_style='medium', color='000000'))
        
        standard_range = ws['B5':'C8']

        for row in standard_range:
            for cell in row:
                cell.border = medium_border
        #######################################################################################
        #기본 셀 셋팅
        ws.merge_cells(start_row = 5, start_column = 3, end_row = 5, end_column = 7)
        ws.merge_cells(start_row = 6, start_column = 3, end_row = 6, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 3, end_row = 7, end_column = 7)
        ws.merge_cells(start_row = 8, start_column = 3, end_row = 8, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 2, end_row = 8, end_column = 2)
        ws['B2'].value = "▶ 연구책임자 : @@@"
        ws['B3'].value = "▶ 연구과제명 :  @@@"
        ws['B5'].value = "회의장소"
        ws['B6'].value = "회의목적"
        ws['B7'].value = "회의내용"
        ws['B11'].value = "참여연구원"
        ws['C10'].value = "이름"
        ws['D10'].value = "소속"

        ws['F11'].value = "외부참석자"
        ws['G10'].value = "이름"
        ws['H10'].value = "소속"
        
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')

        ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H10'].alignment = Alignment(horizontal='center', vertical='center')

        f = Font(name='맑은 고딕', size =10, bold = True)
        ws['B2'].font = f
        ws['B3'].font = f

        #######################################################################################
        

        def check(msg):
            return msg.author == message.author and msg.channel == message.channel
        
        await message.channel.send("네, 알겠습니다. 카드사용일은 언제인가요? ex) 23.10.23")
        while True:
            try:
                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                card_use_date = reply.content
                print(f"카드사용일 : ", card_use_date)
            except asyncio.TimeoutError:
                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                return
            except ValueError:
                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                return

        
            while True:
                await message.channel.send("작성을 완료했습니다. 회의비 금액은 얼마인가요? [숫자로만 작성해주세요.]")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                    meeting_fee = int(reply.content)
                    meeting_fee_str = format(meeting_fee, ',')
                    print(f"회의비 금액 : ", meeting_fee_str)

                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                except ValueError:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    return

                await message.channel.send(f"회의비 금액은  {meeting_fee_str} 원 입니다. 맞습니까? [네/아니오]로 대답")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                if reply.content == "네":
                    await message.channel.send("네, 알겠습니다. 회의 장소는 어디인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response1 = reply.content
                        ws['C5'].value = response1  
                        print("회의 장소 :", ws['C5'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 목적은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response2 = reply.content
                        ws['C6'].value = response2  
                        print("회의 목적 :", ws['C6'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 내용_1은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response3 = reply.content
                        ws['C7'].value = response3  
                        print("회의 내용_1 :", ws['C7'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return


                    await message.channel.send("작성을 완료했습니다. 회의 내용_2는 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response4 = reply.content
                        ws['C8'].value = response4  
                        print("회의 내용_2 :", ws['C8'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return



                    await message.channel.send("작성을 완료했습니다. 내부참석자 중 제외할 사람의 이름을 입력해주세요. [입력을 마치실려면 '끝'이라고 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return

                        if reply.content == "끝":
                            break

                        for person in high_density_list:
                            if person[0] == reply.content:
                                high_density_list.remove(person)
                                break
                        else:
                            await message.channel.send(f"{reply.content}은(는) 내부참석자 목록에 없습니다.")


                    await message.channel.send("작성을 완료했습니다. 내부참석자는 몇명인가요? [숫자로만 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            response5 = int(reply.content)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return
                        if response5 > 7:
                            await message.channel.send("고밀도 내부참석자 인원은 7명 입니다. 다시 입력해주시기 바랍니다.")
                        else:
                            break

                    # 내부참석자 목록을 엑셀에 저장
                    start_row = 11  # 시작할 행 번호
                    start_column = 2

                    print(high_density_list)

                    random_names = random.sample(high_density_list, response5)
                    print("내부참석자 :", random_names)

                    for i, name in enumerate(random_names):
                        cell = ws.cell(row=start_row + i, column=3)  # C열
                        cell.value = name[0]  # 이름 입력
                        cell = ws.cell(row=start_row + i, column=4)  # D열
                        cell.value = name[1]  # 소속 입력

                    #내부 참석자 셀 테두리 칠 범위 구하기
                    internal_participant_range = ws[f'{get_column_letter(start_column)}{start_row - 1}:{get_column_letter(start_column + 2)}{start_row + response5 - 1}']

                    align_cc = Alignment(horizontal='center', vertical='center')

                    for row in internal_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length

                    #---
                    ws.merge_cells(start_row = 11, start_column = 2, end_row = start_row + response5 - 1, end_column = 2)
                    #---

                    external_people = []

                    while True:
                            await message.channel.send("외부참석자 이름을 입력하세요. [존재하지 않을 시 '없음, 없음'이라고 입력해주세요.][입력을 마치실려면 '끝'이라고 입력해주세요]")

                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            
                            if reply.content == "끝":
                                break
                            
                            external_person_name = reply.content
                            
                            await message.channel.send("외부참석자 소속을 입력하세요.")
                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            external_person_affiliation = reply.content
                            
                            external_people.append((external_person_name, external_person_affiliation))

                    
                    for i, person in enumerate(external_people):
                        name_cell = ws.cell(row=start_row + i, column=7)
                        external_people_affiliation_cell = ws.cell(row=start_row + i, column=8)
                        name_cell.value = person[0]
                        external_people_affiliation_cell.value = person[1]



                    ws.merge_cells(start_row = 11, start_column = 6, end_row = start_row + len(external_people) - 1, end_column = 6)



                    external_participant_range = ws[f'{get_column_letter(start_column + 4)}{start_row - 1}:{get_column_letter(start_column + 6)}{start_row + len(external_people) - 1}']

                    for row in external_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length


                    column_widths = [2, 10, 15, 27, 20, 10, 15, 27]


            # 각 열에 대해 너비 지정
                    for column in range(1, len(column_widths) + 1):
                        column_letter = get_column_letter(column)
                        width = column_widths[column - 1]
                        ws.column_dimensions[column_letter].width = width

                    

                    file_name = f"회의비_{meeting_fee_str}.xlsx"
                    file_path = f'C:/Users/ESP/Desktop/Ai_bot/{file_name}'  # 업로드할 파일의 경로
                    zip_file_path = f'C:/Users/ESP/Desktop/Ai_bot/meeting_files{card_use_date}.zip'  # 압축 파일 경로

                    wb.save(file_path)

                    # 파일을 압축하여 zip 파일 생성
                    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                        zipf.write(file_path, file_name)

                    await message.channel.send(f"작성을 완료했습니다. 압축된 엑셀 파일 이름은 '{file_name}'입니다.")

                    file_to_upload = discord.File(zip_file_path, filename=f"high_density_meeting_files_{card_use_date}.zip")  # 압축 파일 업로드
                    await message.channel.send(file=file_to_upload)
                    return
    

                elif reply.content == "아니오":
                    await message.channel.send("다시 질문으로 돌아갑니다.")
                
                    continue
                else:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    continue

    elif message.content == "중견 회의비 작성해줘":
        print("\n\n=========================\n중견 회의비 작성중......")
        print(datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        wb = Workbook()
        ws = wb.active 

        intelligent_list = []


        with open('intelligent_data.csv', 'r', newline='', encoding='EUC-KR') as f:
            reader = csv.reader(f)
            for row in reader:
                intelligent_list.append(row)
        
        print(f"중견 리스트 :" , intelligent_list)

        medium_border = Border(left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            bottom=Side(border_style='medium', color='000000'))
        
        standard_range = ws['B5':'C8']

        for row in standard_range:
            for cell in row:
                cell.border = medium_border
        #######################################################################################
        #기본 셀 셋팅
        ws.merge_cells(start_row = 5, start_column = 3, end_row = 5, end_column = 7)
        ws.merge_cells(start_row = 6, start_column = 3, end_row = 6, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 3, end_row = 7, end_column = 7)
        ws.merge_cells(start_row = 8, start_column = 3, end_row = 8, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 2, end_row = 8, end_column = 2)
        ws['B2'].value = "▶ 연구책임자 : @@@"
        ws['B3'].value = "▶ 연구과제명 : @@@"
        ws['B5'].value = "회의장소"
        ws['B6'].value = "회의목적"
        ws['B7'].value = "회의내용"
        ws['B11'].value = "참여연구원"
        ws['C10'].value = "이름"
        ws['D10'].value = "소속"

        ws['F11'].value = "외부참석자"
        ws['G10'].value = "이름"
        ws['H10'].value = "소속"
        
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')

        ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H10'].alignment = Alignment(horizontal='center', vertical='center')

        f = Font(name='맑은 고딕', size =10, bold = True)
        ws['B2'].font = f
        ws['B3'].font = f

        #######################################################################################
        

        def check(msg):
            return msg.author == message.author and msg.channel == message.channel
        
        await message.channel.send("네, 알겠습니다. 카드사용일은 언제인가요? ex) 23.10.23")
        while True:
            try:
                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                card_use_date = reply.content
                print(f"카드사용일 : ", card_use_date)
            except asyncio.TimeoutError:
                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                return
            except ValueError:
                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                return

        
            while True:
                await message.channel.send("작성을 완료했습니다. 회의비 금액은 얼마인가요? [숫자로만 작성해주세요.]")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                    meeting_fee = int(reply.content)
                    meeting_fee_str = format(meeting_fee, ',')
                    print(f"회의비 금액 : ", meeting_fee_str)

                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                except ValueError:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    return

                await message.channel.send(f"회의비 금액은  {meeting_fee_str} 원 입니다. 맞습니까? [네/아니오]로 대답")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                if reply.content == "네":
                    await message.channel.send("네, 알겠습니다. 회의 장소는 어디인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response1 = reply.content
                        ws['C5'].value = response1  
                        print("회의 장소 :", ws['C5'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 목적은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response2 = reply.content
                        ws['C6'].value = response2  
                        print("회의 목적 :", ws['C6'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 내용_1은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response3 = reply.content
                        ws['C7'].value = response3  
                        print("회의 내용_1 :", ws['C7'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return


                    await message.channel.send("작성을 완료했습니다. 회의 내용_2는 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response4 = reply.content
                        ws['C8'].value = response4  
                        print("회의 내용_2 :", ws['C8'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return



                    await message.channel.send("작성을 완료했습니다. 내부참석자 중 제외할 사람의 이름을 입력해주세요. [입력을 마치실려면 '끝'이라고 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return

                        if reply.content == "끝":
                            break

                        for person in intelligent_list:
                            if person[0] == reply.content:
                                intelligent_list.remove(person)
                                break
                        else:
                            await message.channel.send(f"{reply.content}은(는) 내부참석자 목록에 없습니다.")


                    await message.channel.send("작성을 완료했습니다. 내부참석자는 몇명인가요? [숫자로만 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            response5 = int(reply.content)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return
                        if response5 > 9:
                            await message.channel.send("중견 내부참석자 인원은 9명 입니다. 다시 입력해주시기 바랍니다.")
                        else:
                            break

                    # 내부참석자 목록을 엑셀에 저장
                    start_row = 11  # 시작할 행 번호
                    start_column = 2

                    print(intelligent_list)


                    random_names = random.sample(intelligent_list, response5)
                    print("내부참석자 :", random_names)

                    for i, name in enumerate(random_names):
                        cell = ws.cell(row=start_row + i, column=3)  # C열
                        cell.value = name[0]  # 이름 입력
                        cell = ws.cell(row=start_row + i, column=4)  # D열
                        cell.value = name[1]  # 소속 입력

                    #내부 참석자 셀 테두리 칠 범위 구하기
                    internal_participant_range = ws[f'{get_column_letter(start_column)}{start_row - 1}:{get_column_letter(start_column + 2)}{start_row + response5 - 1}']

                    align_cc = Alignment(horizontal='center', vertical='center')

                    for row in internal_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length

                    #---
                    ws.merge_cells(start_row = 11, start_column = 2, end_row = start_row + response5 - 1, end_column = 2)
                    #---

                    external_people = []

                    while True:
                            await message.channel.send("외부참석자 이름을 입력하세요. [존재하지 않을 시 '없음, 없음'이라고 입력해주세요.][입력을 마치실려면 '끝'이라고 입력해주세요]")

                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            
                            if reply.content == "끝":
                                break
                            
                            external_person_name = reply.content
                            
                            await message.channel.send("외부참석자 소속을 입력하세요.")
                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            external_person_affiliation = reply.content
                            
                            external_people.append((external_person_name, external_person_affiliation))

                    
                    for i, person in enumerate(external_people):
                        name_cell = ws.cell(row=start_row + i, column=7)
                        external_people_affiliation_cell = ws.cell(row=start_row + i, column=8)
                        name_cell.value = person[0]
                        external_people_affiliation_cell.value = person[1]



                    ws.merge_cells(start_row = 11, start_column = 6, end_row = start_row + len(external_people) - 1, end_column = 6)



                    external_participant_range = ws[f'{get_column_letter(start_column + 4)}{start_row - 1}:{get_column_letter(start_column + 6)}{start_row + len(external_people) - 1}']

                    for row in external_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length


                    column_widths = [2, 10, 15, 27, 20, 10, 15, 27]


            # 각 열에 대해 너비 지정
                    for column in range(1, len(column_widths) + 1):
                        column_letter = get_column_letter(column)
                        width = column_widths[column - 1]
                        ws.column_dimensions[column_letter].width = width

                    

                    file_name = f"회의비_{meeting_fee_str}.xlsx"
                    file_path = f'C:/Users/ESP/Desktop/Ai_bot/{file_name}'  # 업로드할 파일의 경로
                    zip_file_path = f'C:/Users/ESP/Desktop/Ai_bot/meeting_files{card_use_date}.zip'  # 압축 파일 경로

                    wb.save(file_path)

                    # 파일을 압축하여 zip 파일 생성
                    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                        zipf.write(file_path, file_name)

                    await message.channel.send(f"작성을 완료했습니다. 압축된 엑셀 파일 이름은 '{file_name}'입니다.")

                    file_to_upload = discord.File(zip_file_path, filename=f"intelligent_meeting_files_{card_use_date}.zip")  # 압축 파일 업로드
                    await message.channel.send(file=file_to_upload)
                    return
    

                elif reply.content == "아니오":
                    await message.channel.send("다시 질문으로 돌아갑니다.")
                
                    continue
                else:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    continue

    elif message.content == "우편물류 회의비 작성해줘":
        print("\n\n=========================\n우편물류 회의비 작성중......")
        print(datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        wb = Workbook()
        ws = wb.active 

        postal_list = []


        with open('postal_data.csv', 'r', newline='', encoding='EUC-KR') as f:
            reader = csv.reader(f)
            for row in reader:
                postal_list.append(row)
        
        print(f"우편물류 리스트 :" , postal_list)

        medium_border = Border(left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            bottom=Side(border_style='medium', color='000000'))
        
        standard_range = ws['B5':'C8']

        for row in standard_range:
            for cell in row:
                cell.border = medium_border
        #######################################################################################
        #기본 셀 셋팅
        ws.merge_cells(start_row = 5, start_column = 3, end_row = 5, end_column = 7)
        ws.merge_cells(start_row = 6, start_column = 3, end_row = 6, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 3, end_row = 7, end_column = 7)
        ws.merge_cells(start_row = 8, start_column = 3, end_row = 8, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 2, end_row = 8, end_column = 2)
        ws['B2'].value = "▶ 연구책임자 : @@@"
        ws['B3'].value = "▶ 연구과제명 :  @@@"
        ws['B5'].value = "회의장소"
        ws['B6'].value = "회의목적"
        ws['B7'].value = "회의내용"
        ws['B11'].value = "참여연구원"
        ws['C10'].value = "이름"
        ws['D10'].value = "소속"

        ws['F11'].value = "외부참석자"
        ws['G10'].value = "이름"
        ws['H10'].value = "소속"
        
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')

        ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H10'].alignment = Alignment(horizontal='center', vertical='center')

        f = Font(name='맑은 고딕', size =10, bold = True)
        ws['B2'].font = f
        ws['B3'].font = f

        #######################################################################################
        

        def check(msg):
            return msg.author == message.author and msg.channel == message.channel
        
        await message.channel.send("네, 알겠습니다. 카드사용일은 언제인가요? ex) 23.10.23")
        while True:
            try:
                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                card_use_date = reply.content
                print(f"카드사용일 : ", card_use_date)
            except asyncio.TimeoutError:
                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                return
            except ValueError:
                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                return

        
            while True:
                await message.channel.send("작성을 완료했습니다. 회의비 금액은 얼마인가요? [숫자로만 작성해주세요.]")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                    meeting_fee = int(reply.content)
                    meeting_fee_str = format(meeting_fee, ',')
                    print(f"회의비 금액 : ", meeting_fee_str)

                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                except ValueError:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    return

                await message.channel.send(f"회의비 금액은  {meeting_fee_str} 원 입니다. 맞습니까? [네/아니오]로 대답")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                if reply.content == "네":
                    await message.channel.send("네, 알겠습니다. 회의 장소는 어디인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response1 = reply.content
                        ws['C5'].value = response1  
                        print("회의 장소 :", ws['C5'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 목적은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response2 = reply.content
                        ws['C6'].value = response2  
                        print("회의 목적 :", ws['C6'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 내용_1은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response3 = reply.content
                        ws['C7'].value = response3  
                        print("회의 내용_1 :", ws['C7'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return


                    await message.channel.send("작성을 완료했습니다. 회의 내용_2는 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response4 = reply.content
                        ws['C8'].value = response4  
                        print("회의 내용_2 :", ws['C8'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return



                    await message.channel.send("작성을 완료했습니다. 내부참석자 중 제외할 사람의 이름을 입력해주세요. [입력을 마치실려면 '끝'이라고 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return

                        if reply.content == "끝":
                            break

                        for person in postal_list:
                            if person[0] == reply.content:
                                postal_list.remove(person)
                                break
                        else:
                            await message.channel.send(f"{reply.content}은(는) 내부참석자 목록에 없습니다.")


                    await message.channel.send("작성을 완료했습니다. 내부참석자는 몇명인가요? [숫자로만 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            response5 = int(reply.content)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return
                        if response5 > 9:
                            await message.channel.send("우편물류 내부참석자 인원은 9명 입니다. 다시 입력해주시기 바랍니다.")
                        else:
                            break

                    # 내부참석자 목록을 엑셀에 저장
                    start_row = 11  # 시작할 행 번호
                    start_column = 2

                    print(postal_list)

                    random_names = random.sample(postal_list, response5)
                    print("내부참석자 :", random_names)

                    for i, name in enumerate(random_names):
                        cell = ws.cell(row=start_row + i, column=3)  # C열
                        cell.value = name[0]  # 이름 입력
                        cell = ws.cell(row=start_row + i, column=4)  # D열
                        cell.value = name[1]  # 소속 입력

                    #내부 참석자 셀 테두리 칠 범위 구하기
                    internal_participant_range = ws[f'{get_column_letter(start_column)}{start_row - 1}:{get_column_letter(start_column + 2)}{start_row + response5 - 1}']

                    align_cc = Alignment(horizontal='center', vertical='center')

                    for row in internal_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length

                    #---
                    ws.merge_cells(start_row = 11, start_column = 2, end_row = start_row + response5 - 1, end_column = 2)
                    #---

                    external_people = []

                    while True:
                            await message.channel.send("외부참석자 이름을 입력하세요. [존재하지 않을 시 '없음, 없음'이라고 입력해주세요.][입력을 마치실려면 '끝'이라고 입력해주세요]")

                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            
                            if reply.content == "끝":
                                break
                            
                            external_person_name = reply.content
                            
                            await message.channel.send("외부참석자 소속을 입력하세요.")
                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            external_person_affiliation = reply.content
                            
                            external_people.append((external_person_name, external_person_affiliation))

                    
                    for i, person in enumerate(external_people):
                        name_cell = ws.cell(row=start_row + i, column=7)
                        external_people_affiliation_cell = ws.cell(row=start_row + i, column=8)
                        name_cell.value = person[0]
                        external_people_affiliation_cell.value = person[1]



                    ws.merge_cells(start_row = 11, start_column = 6, end_row = start_row + len(external_people) - 1, end_column = 6)



                    external_participant_range = ws[f'{get_column_letter(start_column + 4)}{start_row - 1}:{get_column_letter(start_column + 6)}{start_row + len(external_people) - 1}']

                    for row in external_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length


                    column_widths = [2, 10, 15, 27, 20, 10, 15, 27]


            # 각 열에 대해 너비 지정
                    for column in range(1, len(column_widths) + 1):
                        column_letter = get_column_letter(column)
                        width = column_widths[column - 1]
                        ws.column_dimensions[column_letter].width = width

                    

                    file_name = f"회의비_{meeting_fee_str}.xlsx"
                    file_path = f'C:/Users/ESP/Desktop/Ai_bot/{file_name}'  # 업로드할 파일의 경로
                    zip_file_path = f'C:/Users/ESP/Desktop/Ai_bot/meeting_files{card_use_date}.zip'  # 압축 파일 경로

                    wb.save(file_path)

                    # 파일을 압축하여 zip 파일 생성
                    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                        zipf.write(file_path, file_name)

                    await message.channel.send(f"작성을 완료했습니다. 압축된 엑셀 파일 이름은 '{file_name}'입니다.")

                    file_to_upload = discord.File(zip_file_path, filename=f"postal_meeting_files_{card_use_date}.zip")  # 압축 파일 업로드
                    await message.channel.send(file=file_to_upload)
                    return
    

                elif reply.content == "아니오":
                    await message.channel.send("다시 질문으로 돌아갑니다.")
                
                    continue
                else:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    continue
            
    elif message.content == "첨단 회의비 작성해줘":
        print("\n\n=========================\n첨단 DNA 회의비 작성중......")
        print(datetime.today().strftime("%Y-%m-%d %H:%M:%S"))
        wb = Workbook()
        ws = wb.active 

        advanced_DNA_list = []


        with open('advanced_DNA_data.csv', 'r', newline='', encoding='EUC-KR') as f:
            reader = csv.reader(f)
            for row in reader:
                advanced_DNA_list.append(row)
        
        print(f"첨단DNA 참여자 리스트 :" , advanced_DNA_list)

        medium_border = Border(left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            bottom=Side(border_style='medium', color='000000'))
        
        standard_range = ws['B5':'C8']

        for row in standard_range:
            for cell in row:
                cell.border = medium_border
        #######################################################################################
        #기본 셀 셋팅
        ws.merge_cells(start_row = 5, start_column = 3, end_row = 5, end_column = 7)
        ws.merge_cells(start_row = 6, start_column = 3, end_row = 6, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 3, end_row = 7, end_column = 7)
        ws.merge_cells(start_row = 8, start_column = 3, end_row = 8, end_column = 7)
        ws.merge_cells(start_row = 7, start_column = 2, end_row = 8, end_column = 2)
        ws['B2'].value = "▶ 연구책임자 : @@@"
        ws['B3'].value = "▶ 연구과제명 :  @@@"
        ws['B5'].value = "회의장소"
        ws['B6'].value = "회의목적"
        ws['B7'].value = "회의내용"
        ws['B11'].value = "참여연구원"
        ws['C10'].value = "이름"
        ws['D10'].value = "소속"

        ws['F11'].value = "외부참석자"
        ws['G10'].value = "이름"
        ws['H10'].value = "소속"
        
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')

        ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H10'].alignment = Alignment(horizontal='center', vertical='center')

        f = Font(name='맑은 고딕', size =10, bold = True)
        ws['B2'].font = f
        ws['B3'].font = f

        #######################################################################################
        

        def check(msg):
            return msg.author == message.author and msg.channel == message.channel
        
        await message.channel.send("네, 알겠습니다. 카드사용일은 언제인가요? ex) 23.10.23")
        while True:
            try:
                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                card_use_date = reply.content
                print(f"카드사용일 : ", card_use_date)
            except asyncio.TimeoutError:
                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                return
            except ValueError:
                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                return

        
            while True:
                await message.channel.send("작성을 완료했습니다. 회의비 금액은 얼마인가요? [숫자로만 작성해주세요.]")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                    meeting_fee = int(reply.content)
                    meeting_fee_str = format(meeting_fee, ',')
                    print(f"회의비 금액 : ", meeting_fee_str)

                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                except ValueError:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    return

                await message.channel.send(f"회의비 금액은  {meeting_fee_str} 원 입니다. 맞습니까? [네/아니오]로 대답")

                try:
                    reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                except asyncio.TimeoutError:
                    await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                    return
                if reply.content == "네":
                    await message.channel.send("네, 알겠습니다. 회의 장소는 어디인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response1 = reply.content
                        ws['C5'].value = response1  
                        print("회의 장소 :", ws['C5'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 목적은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response2 = reply.content
                        ws['C6'].value = response2  
                        print("회의 목적 :", ws['C6'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return

                    await message.channel.send("작성을 완료했습니다. 회의 내용_1은 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response3 = reply.content
                        ws['C7'].value = response3  
                        print("회의 내용_1 :", ws['C7'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return


                    await message.channel.send("작성을 완료했습니다. 회의 내용_2는 무엇인가요?")

                    try:
                        reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        response4 = reply.content
                        ws['C8'].value = response4  
                        print("회의 내용_2 :", ws['C8'].value)  
                    except asyncio.TimeoutError:
                        await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                        return
                    except ValueError:
                        await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                        return



                    await message.channel.send("작성을 완료했습니다. 내부참석자 중 제외할 사람의 이름을 입력해주세요. [입력을 마치실려면 '끝'이라고 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return

                        if reply.content == "끝":
                            break

                        for person in advanced_DNA_list:
                            if person[0] == reply.content:
                                advanced_DNA_list.remove(person)
                                break
                        else:
                            await message.channel.send(f"{reply.content}은(는) 내부참석자 목록에 없습니다.")


                    await message.channel.send("작성을 완료했습니다. 내부참석자는 몇명인가요? [숫자로만 입력해주세요]")

                    while True:
                        try:
                            reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            response5 = int(reply.content)
                        except asyncio.TimeoutError:
                            await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                            return
                        except ValueError:
                            await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                            return
                        if response5 > 8:
                            await message.channel.send("첨단 DNA 내부참석자 인원은 8명 입니다. 다시 입력해주시기 바랍니다.")
                        else:
                            break

                    # 내부참석자 목록을 엑셀에 저장
                    start_row = 11  # 시작할 행 번호
                    start_column = 2

                    print(advanced_DNA_list)

                    random_names = random.sample(advanced_DNA_list, response5)
                    print("내부참석자 :", random_names)

                    for i, name in enumerate(random_names):
                        cell = ws.cell(row=start_row + i, column=3)  # C열
                        cell.value = name[0]  # 이름 입력
                        cell = ws.cell(row=start_row + i, column=4)  # D열
                        cell.value = name[1]  # 소속 입력

                    #내부 참석자 셀 테두리 칠 범위 구하기
                    internal_participant_range = ws[f'{get_column_letter(start_column)}{start_row - 1}:{get_column_letter(start_column + 2)}{start_row + response5 - 1}']

                    align_cc = Alignment(horizontal='center', vertical='center')

                    for row in internal_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length

                    #---
                    ws.merge_cells(start_row = 11, start_column = 2, end_row = start_row + response5 - 1, end_column = 2)
                    #---

                    external_people = []

                    while True:
                            await message.channel.send("외부참석자 이름을 입력하세요. [존재하지 않을 시 '없음, 없음'이라고 입력해주세요.][입력을 마치실려면 '끝'이라고 입력해주세요]")

                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            
                            if reply.content == "끝":
                                break
                            
                            external_person_name = reply.content
                            
                            await message.channel.send("외부참석자 소속을 입력하세요.")
                            try:
                                reply = await asyncio.wait_for(client.wait_for("message", check=check), timeout=30)
                            except asyncio.TimeoutError:
                                await message.channel.send("시간이 초과되었습니다. 다시 시도해 주세요.")
                                return
                            except ValueError:
                                await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                                return
                            external_person_affiliation = reply.content
                            
                            external_people.append((external_person_name, external_person_affiliation))

                    
                    for i, person in enumerate(external_people):
                        name_cell = ws.cell(row=start_row + i, column=7)
                        external_people_affiliation_cell = ws.cell(row=start_row + i, column=8)
                        name_cell.value = person[0]
                        external_people_affiliation_cell.value = person[1]



                    ws.merge_cells(start_row = 11, start_column = 6, end_row = start_row + len(external_people) - 1, end_column = 6)



                    external_participant_range = ws[f'{get_column_letter(start_column + 4)}{start_row - 1}:{get_column_letter(start_column + 6)}{start_row + len(external_people) - 1}']

                    for row in external_participant_range:
                        for cell in row:
                            cell.border = medium_border
                            cell.alignment = align_cc

                    for column_cells in ws.columns:
                        length = max(len(str(cell.value))*1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length


                    column_widths = [2, 10, 15, 27, 20, 10, 15, 27]


            # 각 열에 대해 너비 지정
                    for column in range(1, len(column_widths) + 1):
                        column_letter = get_column_letter(column)
                        width = column_widths[column - 1]
                        ws.column_dimensions[column_letter].width = width

                    

                    file_name = f"회의비_{meeting_fee_str}.xlsx"
                    file_path = f'C:/Users/ESP/Desktop/Ai_bot/{file_name}'  # 업로드할 파일의 경로
                    zip_file_path = f'C:/Users/ESP/Desktop/Ai_bot/meeting_files{card_use_date}.zip'  # 압축 파일 경로

                    wb.save(file_path)

                    # 파일을 압축하여 zip 파일 생성
                    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                        zipf.write(file_path, file_name)

                    await message.channel.send(f"작성을 완료했습니다. 압축된 엑셀 파일 이름은 '{file_name}'입니다.")

                    file_to_upload = discord.File(zip_file_path, filename=f"advanced_DNA_meeting_files_{card_use_date}.zip")  # 압축 파일 업로드
                    await message.channel.send(file=file_to_upload)
                    return
    

                elif reply.content == "아니오":
                    await message.channel.send("다시 질문으로 돌아갑니다.")
                
                    continue
                else:
                    await message.channel.send("잘못된 입력입니다. 다시 질문으로 돌아갑니다.")
                    continue

client.run(token)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                       